import os
from openpyxl import load_workbook
class riskInfo(object):
    def __init__(self,rsid,rsidUrl,riskA,r):
        self.rsid=rsid
        self.rsidUrl=rsidUrl
        self.riskA=riskA
        self.row_num=r


def getxlsUrl():
    f=os.getcwd()
    aurl="http://grch37.ensembl.org/Homo_sapiens/Variation/Population?db=core;r=8:19808830-19809830;v="
    openFileName=f+"\\pmid.xlsx"
    wb=load_workbook(openFileName)
    sheet1=wb.get_sheet_by_name('gene_snp')
    riskInfoList=[]
    for r in range(1,sheet1.max_row):

        rsid=sheet1['E'][r].value
        riskA=sheet1['U'][r].value
        if riskA:
            riskInfoNew = riskInfo(rsid.strip(),aurl+rsid.strip(),riskA.strip(),r)

            riskInfoList.append(riskInfoNew)

    return  riskInfoList
